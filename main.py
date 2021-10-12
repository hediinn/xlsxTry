from openpyxl import Workbook,load_workbook,chart
from openpyxl.descriptors.base import NoneSet
from openpyxl.reader import excel
from openpyxl.worksheet.cell_range import CellRange
import numpy as np
import matplotlib.pyplot as plt

#This is a class for openpyxl
#This is mostly becouse the dataset is not easy to import and this makes it easier to import diffrent files
#When the class is called the file is loaded in, and when the specific medtod is called the x and y will be set
#In AM02010 the data is beween row 12 and 27
#In AL01020 the data is beween column 3 and 18
class xels:
    def __init__(self,dataset):
        try:
            self.wb =load_workbook(dataset)
        except:
            self.wb = Workbook()
        self.x = []
        self.y = []
        self.ws =self.wb.active

    def printText(self):
        for rol in self.ws.iter_cols(min_row=12, max_col=3, max_row=27):
            for cell in rol:
                if cell.col_idx == 1:
                   self.x.append(cell.value)
                elif cell.col_idx == 3:
                   self.y.append(cell.value)

    def pensPrint(self):
        for rol in self.ws.iter_rows(min_row=3,min_col=3, max_col=18, max_row=5):
            for cell in rol:
                if cell.row == 4:
                       self.x.append(cell.value)
                elif cell.row == 5:
                   self.y.append(cell.value)



e=xels("AM02010.xlsx")
e.printText()
arb=plt.plot(e.x,e.y,'ro-', linewidth=1,label='arbeiðsleys')
d=xels("AL01020 pens.xlsx")
d.pensPrint()
pen=plt.plot(d.x,d.y,'go-', linewidth=1,label='Pensjóistar')
plt.legend()
plt.title('Arbeiðsles í mun til Pensjóistar')
plt.show()
